from flask import request, jsonify, send_file
import os
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity
from database import db
from models.doctor      import Doctor
from models.appointment import Appointment
from models.prescription import Prescription
from models.user        import User
from utils.pdf_generator import generate_prescription_pdf
import json

def _admin_required():
    if str(get_jwt_identity()) != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
    return None

def admin_login():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    admin_email = os.environ.get('ADMIN_EMAIL', '').strip().lower()
    admin_password = os.environ.get('ADMIN_PASSWORD', '')
    if not admin_email or not admin_password:
        return jsonify({'error': 'Admin credentials are not configured'}), 503
    if email != admin_email or password != admin_password:
        return jsonify({'error': 'Invalid admin email or password'}), 401
    token = create_access_token(identity='admin')
    return jsonify({'message': 'Admin login successful', 'token': token}), 200

@jwt_required()
def admin_doctors():
    error = _admin_required()
    if error:
        return error
    doctors = Doctor.query.order_by(Doctor.full_name).all()
    return jsonify({'doctors': [doctor.to_dict() for doctor in doctors]}), 200

@jwt_required()
def admin_add_doctor():
    error = _admin_required()
    if error:
        return error
    data = request.get_json() or {}
    required = ['full_name', 'specialization', 'qualification', 'hospital', 'location', 'email', 'password']
    if any(not str(data.get(field, '')).strip() for field in required):
        return jsonify({'error': 'Full name, specialization, qualification, hospital, location, email, and password are required'}), 400
    email = data['email'].strip().lower()
    if Doctor.query.filter_by(email=email).first():
        return jsonify({'error': 'A doctor with this email already exists'}), 409
    doctor = Doctor(
        full_name=data['full_name'].strip(), specialization=data['specialization'].strip(),
        qualification=data['qualification'].strip(), hospital=data['hospital'].strip(),
        location=data['location'].strip(), email=email,
        experience=data.get('experience', 0), fee=data.get('fee', 500),
        rating=data.get('rating', 4),
        available_days=','.join(data.get('available_days') or ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']),
        slots=','.join(data.get('slots') or ['09:00', '10:00', '11:00', '14:00', '15:00']),
        is_active=bool(data.get('is_active', True)),
        reg_number=data.get('reg_number', '').strip() or None,
    )
    doctor.set_password(data['password'])
    db.session.add(doctor)
    db.session.commit()
    return jsonify({'message': 'Doctor added successfully', 'doctor': doctor.to_dict()}), 201

@jwt_required()
def admin_update_status(doctor_id):
    error = _admin_required()
    if error:
        return error
    doctor = Doctor.query.get(doctor_id)
    if not doctor:
        return jsonify({'error': 'Doctor not found'}), 404
    data = request.get_json() or {}
    if 'is_active' not in data:
        return jsonify({'error': 'is_active is required'}), 400
    doctor.is_active = bool(data['is_active'])
    db.session.commit()
    return jsonify({'message': 'Doctor status updated', 'doctor': doctor.to_dict()}), 200

def doctor_login():
    data     = request.get_json()
    email    = data.get('email','').strip().lower()
    password = data.get('password','')
    doctor   = Doctor.query.filter_by(email=email).first()
    if not doctor or not doctor.check_password(password):
        return jsonify({'error':'Invalid email or password'}), 401
    token = create_access_token(identity=f"doctor_{doctor.id}")
    return jsonify({'message':'Login successful','token':token,'doctor':doctor.to_dict()}), 200

@jwt_required()
def doctor_profile():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor = Doctor.query.get(int(str(identity).replace('doctor_','')))
    if not doctor:
        return jsonify({'error':'Doctor not found'}), 404
    return jsonify({'doctor':doctor.to_dict()}), 200

@jwt_required()
def update_doctor_profile():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor = Doctor.query.get(int(str(identity).replace('doctor_','')))
    if not doctor:
        return jsonify({'error':'Doctor not found'}), 404
    data   = request.get_json()
    fields = ['full_name','age','gender','mobile','bio','languages',
              'awards','publications','consultation_type','photo_url',
              'signature_text','hospital','location','fee','qualification']
    for f in fields:
        if f in data:
            setattr(doctor, f, data[f])
    db.session.commit()
    return jsonify({'message':'Profile updated','doctor':doctor.to_dict()}), 200

@jwt_required()
def my_patients():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor_id = int(str(identity).replace('doctor_',''))
    appts = Appointment.query.filter_by(doctor_id=doctor_id)        .filter(Appointment.status != 'CANCELLED')        .order_by(Appointment.date, Appointment.time_slot).all()
    result = []
    for a in appts:
        patient = User.query.get(a.user_id)
        d = a.to_dict()
        d['patient'] = patient.to_dict() if patient else None
        result.append(d)
    return jsonify({'appointments':result,'count':len(result)}), 200

@jwt_required()
def write_prescription():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor_id = int(str(identity).replace('doctor_',''))
    data = request.get_json()
    appointment_id = data.get('appointment_id')
    diagnosis      = data.get('diagnosis','')
    medicines      = data.get('medicines',[])
    advice         = data.get('advice','')
    follow_up_date = data.get('follow_up_date','')

    appt = Appointment.query.filter_by(id=appointment_id, doctor_id=doctor_id).first()
    if not appt:
        return jsonify({'error':'Appointment not found'}), 404

    existing = Prescription.query.filter_by(appointment_id=appointment_id).first()
    if existing:
        existing.diagnosis      = diagnosis
        existing.medicines      = json.dumps(medicines)
        existing.advice         = advice
        existing.follow_up_date = follow_up_date
        pres = existing
    else:
        pres = Prescription(
            appointment_id=appointment_id, user_id=appt.user_id,
            doctor_id=doctor_id, diagnosis=diagnosis,
            medicines=json.dumps(medicines), advice=advice,
            follow_up_date=follow_up_date,
        )
        db.session.add(pres)

    appt.status = 'COMPLETED'
    db.session.flush()

    doctor  = Doctor.query.get(doctor_id)
    patient = User.query.get(appt.user_id)
    try:
        filename, filepath = generate_prescription_pdf(pres, doctor, patient, appt)
        pres.pdf_path = filename
    except Exception as e:
        print(f"PDF error: {e}")
        pres.pdf_path = None

    db.session.commit()
    return jsonify({'message':'Prescription saved','prescription':pres.to_dict()}), 201

@jwt_required()
def download_prescription(prescription_id):
    identity = get_jwt_identity()
    pres     = Prescription.query.get(prescription_id)
    if not pres:
        return jsonify({'error':'Prescription not found'}), 404

    uid = str(identity)
    is_doctor  = uid.startswith('doctor_') and int(uid.replace('doctor_','')) == pres.doctor_id
    is_patient = not uid.startswith('doctor_') and int(uid) == pres.user_id
    if not is_doctor and not is_patient:
        return jsonify({'error':'Unauthorized'}), 403

    if not pres.pdf_path:
        return jsonify({'error':'PDF not generated yet'}), 404

    pdf_dir  = os.path.join(os.path.dirname(__file__), '..', 'pdfs')
    filepath = os.path.join(pdf_dir, pres.pdf_path)
    if not os.path.exists(filepath):
        return jsonify({'error':'PDF file not found on server'}), 404

    return send_file(filepath, as_attachment=True,
                     download_name=f"prescription_{prescription_id}.pdf")

@jwt_required()
def update_availability():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor_id = int(str(identity).replace('doctor_',''))
    doctor    = Doctor.query.get(doctor_id)
    data      = request.get_json()
    if 'available_days' in data:
        doctor.available_days = ','.join(data['available_days'])
    if 'slots' in data:
        doctor.slots = ','.join(data['slots'])
    db.session.commit()
    return jsonify({'message':'Availability updated','doctor':doctor.to_dict()}), 200

@jwt_required()
def export_appointments():
    identity = get_jwt_identity()
    if not str(identity).startswith('doctor_'):
        return jsonify({'error':'Unauthorized'}), 403
    doctor_id = int(str(identity).replace('doctor_',''))
    appts = Appointment.query.filter_by(doctor_id=doctor_id).all()
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"
        headers = ['ID','Patient','Date','Time','Symptoms','Severity','Status']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='0F6E56')
            cell.alignment = Alignment(horizontal='center')
        for row_num, a in enumerate(appts, 2):
            patient = User.query.get(a.user_id)
            ws.append([
                a.id, patient.full_name if patient else 'Unknown',
                a.date, a.time_slot, a.symptoms or '',
                a.severity, a.status,
            ])
        import io
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                        download_name='appointments.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        return jsonify({'error':'openpyxl not installed'}), 500